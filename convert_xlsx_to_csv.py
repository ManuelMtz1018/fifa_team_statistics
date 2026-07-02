import csv
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta instalar la dependencia: openpyxl")
    print("Ejecuta: pip install -r requirements.txt")
    sys.exit(1)


INPUT_ENV = "INPUT_XLSX_PATHS"
OUTPUT_ENV = "OUTPUT_CSV_DIR"
CONVERT_ALL_SHEETS_ENV = "CONVERT_ALL_SHEETS"
ENV_FILE = ".env"


def load_env_file(env_file=ENV_FILE):
    env_path = Path(env_file)
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")

        if key and key not in os.environ:
            os.environ[key] = value


def truthy(value):
    return str(value).strip().lower() in {"1", "true", "yes", "y", "si", "sí"}


def safe_filename(value):
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip()
    return cleaned or "sheet"


def discover_xlsx_files(input_value):
    files = []

    for raw_item in input_value.split(";"):
        item = raw_item.strip().strip('"').strip("'")
        if not item:
            continue

        path = Path(item)

        if any(char in item for char in ["*", "?", "["]):
            files.extend(Path().glob(item) if not path.is_absolute() else path.parent.glob(path.name))
            continue

        if path.is_dir():
            files.extend(path.glob("*.xlsx"))
            continue

        files.append(path)

    unique_files = []
    seen = set()
    for file_path in files:
        resolved = file_path.resolve()
        if resolved not in seen and file_path.suffix.lower() == ".xlsx":
            seen.add(resolved)
            unique_files.append(resolved)

    return sorted(unique_files)


def cell_to_csv_value(cell_value):
    if cell_value is None:
        return ""
    return cell_value


def worksheet_has_data(worksheet):
    for row in worksheet.iter_rows(values_only=True):
        if any(value is not None for value in row):
            return True
    return False


def write_worksheet_to_csv(worksheet, output_file):
    with output_file.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow([cell_to_csv_value(value) for value in row])


def convert_workbook(xlsx_file, output_dir, convert_all_sheets):
    workbook = load_workbook(xlsx_file, read_only=True, data_only=True)

    try:
        worksheets = workbook.worksheets if convert_all_sheets else [workbook.active]
        exported = []

        for worksheet in worksheets:
            if convert_all_sheets:
                csv_name = f"{safe_filename(xlsx_file.stem)}__{safe_filename(worksheet.title)}.csv"
            else:
                csv_name = f"{safe_filename(xlsx_file.stem)}.csv"

            output_file = output_dir / csv_name
            write_worksheet_to_csv(worksheet, output_file)

            exported.append(
                {
                    "sheet": worksheet.title,
                    "output": output_file,
                    "empty": not worksheet_has_data(worksheet),
                }
            )

        return exported
    finally:
        workbook.close()


def start_converter_process()-> None:
    load_env_file()

    input_value = os.environ.get(INPUT_ENV, "").strip()
    output_value = os.environ.get(OUTPUT_ENV, "").strip()
    convert_all_sheets = truthy(os.environ.get(CONVERT_ALL_SHEETS_ENV, "false"))

    if not input_value:
        print(f"Configura la variable {INPUT_ENV} en el archivo .env.")
        print(r"Ejemplo: INPUT_XLSX_PATHS=C:\datos\archivo1.xlsx;C:\datos\archivo2.xlsx")
        return 2

    if not output_value:
        print(f"Configura la variable {OUTPUT_ENV} en el archivo .env.")
        print(r"Ejemplo: OUTPUT_CSV_DIR=C:\datos\csv")
        return 2

    output_dir = Path(output_value).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = discover_xlsx_files(input_value)
    if not xlsx_files:
        print("No se encontraron archivos .xlsx para convertir.")
        return 1

    total_exports = 0
    failures = 0

    for xlsx_file in xlsx_files:
        if not xlsx_file.exists():
            print(f"No existe: {xlsx_file}")
            failures += 1
            continue

        try:
            exports = convert_workbook(xlsx_file, output_dir, convert_all_sheets)
            total_exports += len(exports)
            for export in exports:
                empty_note = " (hoja vacia)" if export["empty"] else ""
                print(f"OK: {xlsx_file.name} [{export['sheet']}] -> {export['output']}{empty_note}")
        except Exception as exc:
            print(f"ERROR: {xlsx_file} -> {exc}")
            failures += 1

    print(f"Conversion terminada. CSV generados: {total_exports}. Errores: {failures}.")
    return 1 if failures else 0


